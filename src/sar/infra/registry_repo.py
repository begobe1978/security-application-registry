# Copyright (C) 2026 Bernardo Gómez Bey
# SPDX-License-Identifier: AGPL-3.0-or-later

from __future__ import annotations

import hashlib
import json
import shutil
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from sar.core.utils import canon, normalize_columns


def _norm_key(s: str) -> str:
    """Normalise a header/field key to a stable snake_case-like lower format."""
    return str(s or "").strip().replace(" ", "_").replace("-", "_").lower()


def update_fields_existing(path: str, sheet: str, human_id: str, fields: dict[str, object]) -> None:
    """Update one or more existing columns for a row identified by human_id.

    - Only updates columns that already exist in the sheet header.
    - Uses openpyxl to preserve formatting.
    """
    if not fields:
        return

    wb = load_workbook(path)
    if sheet not in wb.sheetnames:
        raise ValueError(f"No existe la pestaña '{sheet}' en el Excel.")

    ws = wb[sheet]

    # Header mapping (normalised header -> column index)
    headers: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v is None:
            continue
        headers[_norm_key(v)] = col

    if "human_id" not in headers:
        raise ValueError(f"La pestaña '{sheet}' no tiene columna 'human_id'.")

    # Validate fields exist
    missing = [k for k in fields.keys() if _norm_key(k) not in headers]
    if missing:
        missing_str = ", ".join(missing)
        raise ValueError(
            f"Columnas inexistentes en '{sheet}': {missing_str}. "
            "(En este paso solo se permite actualizar columnas ya existentes.)"
        )

    # Find row
    col_hid = headers["human_id"]
    target = canon(human_id)
    found_row = None
    for row in range(2, ws.max_row + 1):
        v = ws.cell(row=row, column=col_hid).value
        if canon(str(v or "")) == target:
            found_row = row
            break

    if not found_row:
        raise ValueError(f"No se encontró '{human_id}' en '{sheet}'.")

    # Write updates
    for key, value in fields.items():
        col_idx = headers[_norm_key(key)]
        ws.cell(row=found_row, column=col_idx).value = value

    wb.save(path)


def add_new_field_column(path: str, sheet: str, human_id: str, field_name: str, value: object) -> None:
    """Add a brand-new column to a sheet and set its value for the given human_id.

    - The column must NOT already exist (checked by canon/normalisation).
    - Uses openpyxl to preserve formatting.
    """
    field_name = str(field_name or "").strip()
    if not field_name:
        raise ValueError("El nombre del campo no puede estar vacío.")

    wb = load_workbook(path)
    if sheet not in wb.sheetnames:
        raise ValueError(f"No existe la pestaña '{sheet}' en el Excel.")

    ws = wb[sheet]

    # Build header mapping (normalised header -> column index) and also keep original headers.
    headers: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v is None:
            continue
        headers[_norm_key(v)] = col

    if "human_id" not in headers:
        raise ValueError(f"La pestaña '{sheet}' no tiene columna 'human_id'.")

    key = _norm_key(field_name)
    if key in headers:
        raise ValueError(f"El campo '{field_name}' ya existe en '{sheet}'.")

    # Find row
    col_hid = headers["human_id"]
    target = canon(human_id)
    found_row = None
    for row in range(2, ws.max_row + 1):
        v = ws.cell(row=row, column=col_hid).value
        if canon(str(v or "")) == target:
            found_row = row
            break

    if not found_row:
        raise ValueError(f"No se encontró '{human_id}' en '{sheet}'.")

    # Append new column at the end
    new_col = ws.max_column + 1
    ws.cell(row=1, column=new_col).value = field_name
    ws.cell(row=found_row, column=new_col).value = value

    wb.save(path)




def generate_next_human_id(path: str, sheet: str, prefix: str) -> str:
    """Generate the next sequential human_id for a given sheet/prefix.

    This scans existing human_id values and returns PREFIX-### using 3-digit padding.
    """
    wb = load_workbook(path, read_only=True)
    if sheet not in wb.sheetnames:
        raise ValueError(f"No existe la pestaña '{sheet}' en el Excel.")
    ws = wb[sheet]

    # Header mapping
    headers: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v is None:
            continue
        headers[_norm_key(v)] = col
    if "human_id" not in headers:
        raise ValueError(f"La pestaña '{sheet}' no tiene columna 'human_id'.")

    col_hid = headers["human_id"]
    max_n = 0
    pref = canon(prefix)
    for row in range(2, ws.max_row + 1):
        v = ws.cell(row=row, column=col_hid).value
        s = canon(str(v or ""))
        if not s.startswith(pref):
            continue
        m = __import__("re").match(r"^" + __import__("re").escape(pref) + r"(\d+)$", s)
        if m:
            try:
                max_n = max(max_n, int(m.group(1)))
            except Exception:
                pass

    return f"{prefix}{max_n + 1:03d}"


def append_row_existing_columns(path: str, sheet: str, row: dict[str, object]) -> None:
    """Append a new row to a sheet, writing only existing columns.

    - Uses openpyxl to preserve formatting.
    - Does NOT create new columns.
    """
    wb = load_workbook(path)
    if sheet not in wb.sheetnames:
        raise ValueError(f"No existe la pestaña '{sheet}' en el Excel.")
    ws = wb[sheet]

    # Header mapping
    headers: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v is None:
            continue
        headers[_norm_key(v)] = col

    if "human_id" not in headers:
        raise ValueError(f"La pestaña '{sheet}' no tiene columna 'human_id'.")

    # Find last data row by scanning human_id column (avoid trailing formatted rows)
    col_hid = headers["human_id"]
    last = 1
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=col_hid).value
        if str(v or "").strip():
            last = r
    new_row = last + 1

    # Ensure unique human_id
    new_hid = canon(str(row.get("human_id", "")))
    if not new_hid:
        raise ValueError("El registro nuevo debe incluir 'human_id'.")
    for r in range(2, last + 1):
        v = ws.cell(row=r, column=col_hid).value
        if canon(str(v or "")) == new_hid:
            raise ValueError(f"Ya existe un registro con human_id '{row.get('human_id')}' en '{sheet}'.")

    # Write values for existing columns only
    for key, value in row.items():
        k = _norm_key(key)
        if k not in headers:
            continue
        ws.cell(row=new_row, column=headers[k]).value = value

    wb.save(path)


def read_sheet(path: str, sheet: str) -> pd.DataFrame:
    """Read a sheet from the registry Excel into a normalised dataframe."""
    df = pd.read_excel(path, sheet_name=sheet, dtype=str).fillna("")
    return normalize_columns(df)


def read_meta_dict(path: str) -> dict[str, str]:
    """Read META sheet as a key/value dictionary (both as strings)."""
    try:
        df = pd.read_excel(path, sheet_name="META", dtype=str).fillna("")
    except Exception:
        return {}
    df = normalize_columns(df)
    if "key" not in df.columns or "value" not in df.columns:
        return {}
    out: dict[str, str] = {}
    for _, r in df.iterrows():
        k = str(r.get("key", "") or "").strip()
        if not k:
            continue
        out[k] = str(r.get("value", "") or "")
    return out


def write_meta_kv(path: str, updates: dict[str, str]) -> None:
    """Upsert key/value pairs in META (preserving formatting with openpyxl)."""
    if not updates:
        return
    wb = load_workbook(path)
    if "META" not in wb.sheetnames:
        raise ValueError("No existe la pestaña 'META' en el Excel.")
    ws = wb["META"]

    # Find key/value columns by header
    headers: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is None:
            continue
        headers[_norm_key(v)] = c
    if "key" not in headers or "value" not in headers:
        raise ValueError("La pestaña 'META' debe contener columnas 'key' y 'value'.")
    c_key = headers["key"]
    c_val = headers["value"]

    # Build existing row index
    key_to_row: dict[str, int] = {}
    for r in range(2, ws.max_row + 1):
        k = str(ws.cell(row=r, column=c_key).value or "").strip()
        if k:
            key_to_row[k] = r

    # Append/update
    last = ws.max_row
    for k, v in updates.items():
        ks = str(k or "").strip()
        if not ks:
            continue
        if ks in key_to_row:
            rr = key_to_row[ks]
        else:
            last += 1
            rr = last
            ws.cell(row=rr, column=c_key).value = ks
        ws.cell(row=rr, column=c_val).value = str(v or "")

    wb.save(path)


def get_sheet_headers(path: str, sheet: str) -> list[str]:
    """Return normalised headers for a given sheet (row 1), excluding empty headers."""
    wb = load_workbook(path, read_only=True)
    if sheet not in wb.sheetnames:
        return []
    ws = wb[sheet]
    out: list[str] = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is None:
            continue
        s = str(v).strip()
        if not s:
            continue
        out.append(_norm_key(s))
    # De-dup while keeping order
    seen = set()
    uniq = []
    for h in out:
        if h in seen:
            continue
        seen.add(h)
        uniq.append(h)
    return uniq


def get_schema_map(path: str, sheets: list[str]) -> dict[str, list[str]]:
    """Return schema map: {sheet_name: [normalised_header, ...]}"""
    sm: dict[str, list[str]] = {}
    for sh in sheets:
        sm[sh] = get_sheet_headers(path, sh)
    return sm


def schema_hash(schema_map: dict[str, list[str]]) -> str:
    """Stable SHA1 hash for a schema map (sorted by sheet name)."""
    payload = {k: schema_map[k] for k in sorted(schema_map.keys())}
    raw = json.dumps(payload, ensure_ascii=False, separators=(",", ":"), sort_keys=True)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()


def add_missing_columns(path: str, sheet: str, columns: list[str]) -> None:
    """Add columns (headers only) to a sheet if missing. Columns are given as *raw* header names."""
    cols = [str(c or "").strip() for c in (columns or []) if str(c or "").strip()]
    if not cols:
        return
    wb = load_workbook(path)
    if sheet not in wb.sheetnames:
        raise ValueError(f"No existe la pestaña '{sheet}' en el Excel.")
    ws = wb[sheet]

    existing = get_sheet_headers(path, sheet)
    existing_set = set(existing)

    # Append at end. We write the raw header exactly as passed.
    for raw in cols:
        nk = _norm_key(raw)
        if nk in existing_set:
            continue
        new_col = ws.max_column + 1
        ws.cell(row=1, column=new_col).value = raw
        existing_set.add(nk)

    wb.save(path)


def read_lookups(path: str) -> pd.DataFrame:
    """Read LOOKUPS sheet into a normalised dataframe.

    Expected columns (after normalisation):
      - lookup_name
      - lookup_value
      - level (e.g. ALL / C1 / C2 / C3 / C4)
      - description (optional)
    """
    df = pd.read_excel(path, sheet_name="LOOKUPS", dtype=str).fillna("")
    return normalize_columns(df)


def lookup_options_by_level(path: str, level: str) -> dict[str, list[dict[str, str]]]:
    """Return dropdown options for a given level, keyed by field name.

    Convention: lookup_name == field name.
    """
    lvl = str(level or "").strip().upper()
    df = read_lookups(path)
    if df.empty:
        base: dict[str, list[dict[str, str]]] = {}
    else:
        df["level"] = df.get("level", "").astype(str).str.upper().str.strip()
        df["lookup_name"] = df.get("lookup_name", "").astype(str).str.strip()
        df["lookup_value"] = df.get("lookup_value", "").astype(str).str.strip()
        df["description"] = df.get("description", "").astype(str).str.strip()

        scoped = df[(df["level"].isin(["ALL", lvl])) & (df["lookup_name"] != "") & (df["lookup_value"] != "")]

        base = {}
        for name, g in scoped.groupby("lookup_name"):
            opts = []
            for _, r in g.iterrows():
                opts.append({"value": r["lookup_value"], "label": r.get("description", "") or r["lookup_value"]})
            base[str(name)] = opts

    # Backwards-compatible aliases / common column naming variants.
    # Example: registry uses 'environments' column while LOOKUPS may define 'environment'.
    if "environment" in base and "environments" not in base:
        base["environments"] = list(base["environment"])
    if "environments" in base and "environment" not in base:
        base["environment"] = list(base["environments"])

    # Hard-coded safety enums not necessarily present in LOOKUPS
    if "vulnerabilities_detected" not in base:
        base["vulnerabilities_detected"] = [
            {"value": "yes", "label": "yes"},
            {"value": "no", "label": "no"},
            {"value": "unknown", "label": "unknown"},
        ]

    return base


def backup_registry(path: str) -> str:
    """Create a timestamped .bak copy next to the registry file."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    src = Path(path)
    dst = src.with_suffix(src.suffix + f".bak_{ts}")
    shutil.copy2(src, dst)
    return str(dst)


def set_status(path: str, sheet: str, human_id: str, new_status: str) -> None:
    """Update status cell for a row identified by human_id.

    Uses openpyxl to preserve formatting.
    """
    wb = load_workbook(path)
    if sheet not in wb.sheetnames:
        raise ValueError(f"No existe la pestaña '{sheet}' en el Excel.")

    ws = wb[sheet]

    headers = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v is None:
            continue
        key = str(v).strip().replace(" ", "_").replace("-", "_").lower()
        headers[key] = col

    if "human_id" not in headers:
        raise ValueError(f"La pestaña '{sheet}' no tiene columna 'human_id'.")
    if "status" not in headers:
        raise ValueError(f"La pestaña '{sheet}' no tiene columna 'status'.")

    col_hid = headers["human_id"]
    col_status = headers["status"]

    target = canon(human_id)
    found_row = None
    for row in range(2, ws.max_row + 1):
        v = ws.cell(row=row, column=col_hid).value
        if canon(str(v or "")) == target:
            found_row = row
            break

    if not found_row:
        raise ValueError(f"No se encontró '{human_id}' en '{sheet}'.")

    ws.cell(row=found_row, column=col_status).value = new_status
    wb.save(path)
