import os
import sys
from pathlib import Path as _Path
sys.path.insert(0, str(_Path(__file__).resolve().parents[1] / "src"))

import shutil
from pathlib import Path

import pytest
from openpyxl import load_workbook


@pytest.fixture()
def tmp_registry(tmp_path: Path) -> Path:
    """
    Create a minimal, valid registry.xlsx in a temporary directory based on the shipped template.
    Includes:
      - 1 chain: PRJ-0001 -> APP-0001 -> CMP-0001 -> RUN-0001
      - 1 C3 sibling: CMP-0002 (same APP)
      - 1 C4 sibling: RUN-0002 (same CMP-0001)
    """
    # Use the repository template as base
    repo_root = Path(__file__).resolve().parents[1]
    template_path = repo_root / "data" / "registry_template.xlsx"
    assert template_path.exists(), f"Missing template: {template_path}"

    data_dir = tmp_path / "data"
    data_dir.mkdir(parents=True, exist_ok=True)

    reg_path = data_dir / "registry.xlsx"

    shutil.copyfile(template_path, reg_path)
    # Also keep a copy named registry_template.xlsx for app configuration/tests
    shutil.copyfile(template_path, data_dir / "registry_template.xlsx")

    wb = load_workbook(reg_path)

    def write_row(sheet_name: str, values_by_col: dict):
        ws = wb[sheet_name]
        headers = [c.value for c in ws[1]]
        row = [None] * len(headers)
        for col, val in values_by_col.items():
            idx = headers.index(col)
            row[idx] = val
        ws.append(row)

    # Minimal chain
    write_row("C1_Proyectos", {"human_id": "PRJ-0001", "status": "active", "name": "Proyecto Uno"})
    write_row(
        "C2_Aplicaciones",
        {"c1_human_id": "PRJ-0001", "human_id": "APP-0001", "status": "active", "name": "App Uno"},
    )
    write_row(
        "C3_Componentes",
        {"c2_human_id": "APP-0001", "human_id": "CMP-0001", "status": "active", "name": "Componente Uno"},
    )
    write_row(
        "C4_Runtime",
        {"c3_human_id": "CMP-0001", "human_id": "RUN-0001", "status": "active", "name": "Runtime Uno"},
    )

    # Siblings for context lists
    write_row(
        "C3_Componentes",
        {"c2_human_id": "APP-0001", "human_id": "CMP-0002", "status": "active", "name": "Componente Dos"},
    )
    write_row(
        "C4_Runtime",
        {"c3_human_id": "CMP-0001", "human_id": "RUN-0002", "status": "active", "name": "Runtime Dos"},
    )

    wb.save(reg_path)
    return reg_path


@pytest.fixture()
def report_template_path() -> Path:
    repo_root = Path(__file__).resolve().parents[1]
    p = repo_root / "src" / "sar" / "report_templates" / "c4_chain_report.docx"
    assert p.exists(), f"Missing report template: {p}"
    return p


@pytest.fixture()
def no_mmdc(monkeypatch):
    # Ensure we test the fallback path (no Mermaid CLI available)
    monkeypatch.delenv("SAR_MMDC_PATH", raising=False)
    # Also ensure "mmdc" can't be found by setting PATH empty-ish
    monkeypatch.setenv("PATH", "")
